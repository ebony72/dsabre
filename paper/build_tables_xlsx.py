"""Collect all tables from dsabre.tex and appendices.tex into a single .xlsx."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

bold = Font(bold=True)
header_fill = PatternFill("solid", fgColor="D9E1F2")
title_font = Font(bold=True, size=12)

def add_sheet(name, title, headers, rows, notes=None):
    ws = wb.create_sheet(name[:31])
    ws["A1"] = title
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 2))
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for i, row in enumerate(rows, 4):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)
    if notes:
        ws.cell(row=4 + len(rows) + 1, column=1, value="Notes: " + notes)
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.column_dimensions["A"].width = 22

# ── Index sheet ──
idx = wb.create_sheet("Index")
idx["A1"] = "dSABRE paper — all tables"
idx["A1"].font = title_font
idx_rows = [
    ("Sheet", "Table label", "Caption"),
    ("T_example", "tab:example", "Score breakdown for the three teleport candidates of the running example"),
    ("T_circuits", "tab:circuits", "Benchmark circuit properties"),
    ("T_params", "tab:params", "Hardware cost and heuristic parameters"),
    ("T_main25", "tab:main25", "25-qubit suite (B-grid 2x2 4x4)"),
    ("T_main36", "tab:main36", "36-qubit suite (B-grid 2x2 4x4)"),
    ("T_main64", "tab:main64", "64-qubit suite (H-grid 2x3 4x4)"),
    ("T_ablation", "tab:ablation", "Extended-set ablation: topological vs BFS"),
    ("T_mech", "tab:mech", "Mechanism ablation (25q and 64q)"),
    ("T_timing", "tab:timing", "Compilation time, 64q suite"),
    ("T_large", "tab:large", "Large-circuit scalability (QFT 100q/200q/360q)"),
    ("T_sensitivity", "tab:sensitivity", "Cost-model sensitivity (cost_tele sweep)"),
    ("T_nodedecay_25", "tab:node_decay_25", "Node-decay ablation, 25q"),
    ("T_nodedecay_36", "tab:node_decay_36", "Node-decay ablation, 36q"),
    ("T_nodedecay_64", "tab:node_decay_64", "Node-decay ablation, 64q"),
    ("T_hopgain", "tab:hop_gain", "Hop-gain ablation, all three suites"),
    ("T_link_density", "tab:link_density", "4-link vs 8-link B-grid"),
    ("T_layout", "tab:layout", "Initial-layout method x routing schedule (gmean EPR)"),
    ("T_pytket_25_36", "tab:pytket_shared_25_36", "Shared KaHyPar mapping: pytket-dqc vs dSABRE (25q,36q)"),
    ("T_pytket_64", "tab:pytket_shared_64", "Shared KaHyPar mapping: pytket-dqc vs dSABRE (64q)"),
    ("T_dmaps_results", "tab:dmaps_results", "DMapS detailed results (best/mean/stdev)"),
    ("T_dmaps_vs_dsabre", "tab:dmaps_vs_dsabre", "DMapS vs dSABRE head-to-head"),
    ("T_dmaps_fill", "tab:dmaps_fill", "Fill-ratio sweep (QFT on fixed 64-phys B-grid)"),
]
for i, row in enumerate(idx_rows, 3):
    for j, v in enumerate(row, 1):
        c = idx.cell(row=i, column=j, value=v)
        if i == 3:
            c.font = bold
            c.fill = header_fill
idx.column_dimensions["A"].width = 22
idx.column_dimensions["B"].width = 26
idx.column_dimensions["C"].width = 70

# ── tab:example ──
add_sheet("T_example", "Table: Score breakdown — running example (Fig. tele-candidates)",
    ["Candidate", "next core", "direction", "d_prep", "c_cap", "g_hop", "score s"],
    [["A", "C2", "towards C5", 1, 0, 5, -4],
     ["B", "C0", "away from C5", 2, 0, -5, 7],
     ["C", "C4", "towards C5", 3, 15, 5, 13]],
    notes="Lower is better; Candidate A is selected.")

# ── tab:circuits ──
add_sheet("T_circuits", "Table: Benchmark circuit properties (measurements/barriers stripped)",
    ["Suite", "Circuit", "Qubits", "CX gates", "Depth", "CX/qubit"],
    [["25q","AE",25,558,395,22.3],["25q","QFT",25,580,173,23.2],["25q","QNN",25,1223,259,48.9],
     ["25q","Random",25,1124,589,45.0],["25q","GHZ",25,24,27,1.0],["25q","Graphstate",25,25,19,1.0],
     ["36q","BV",36,17,23,0.5],["36q","DJ",36,35,41,1.0],["36q","W-state",36,70,145,1.9],
     ["36q","VQE-SU2",36,105,56,2.9],["36q","QPEexact",36,1019,347,28.3],["36q","QAOA",36,1200,256,33.3],
     ["64q","AE",64,1962,1058,30.7],["64q","QFT",64,1966,446,30.7],["64q","QNN",64,8126,650,127.0],
     ["64q","Random",64,1627,403,25.4],["64q","GHZ",64,63,66,1.0],["64q","Graphstate",64,64,25,1.0]])

# ── tab:params ──
add_sheet("T_params", "Table: Hardware cost and heuristic parameters",
    ["Parameter", "Symbol", "Value"],
    [["SWAP cost","c_swap",3],["Teleport (EPR) cost","c_tele",10],
     ["Capacity threshold","tau",3],["Capacity penalty","c_pen",15],
     ["Hop-gain weight","w_h",5],["Extended-set weight","w_e",0.25],
     ["Extended-set capacity","L",20],["Lookahead decay","gamma",0.9],
     ["Deadlock limit","Delta_lim",50],["Max rollbacks","N_bkp",50]])

# ── tab:main25 ──
add_sheet("T_main25", "Table: 25-qubit suite (B-grid 2x2 4x4, 64 physical qubits)",
    ["Circuit","CX","TS EPR","TS SWAP","dSABRE EPR","dSABRE SWAP","tket e-bits","Δ vs TS (%)","Δ vs tket (%)"],
    [["ae",558,23,297,23,374,85,0.0,-72.9],
     ["ghz",24,2,29,1,14,3,-50.0,-66.7],
     ["graphstate",25,11,51,2,16,4,-81.8,-50.0],
     ["qft",580,39,355,33,461,120,-15.4,-72.5],
     ["qnn",1223,51,587,48,1076,152,-5.9,-68.4],
     ["random",1124,292,1273,169,1374,665,-42.1,-74.6],
     ["gmean","",25.8,221,15.2,196,48.1,-41.1,-68.4]],
    notes="TS best-of-3 seed; pytket-dqc best-of-5 seed; dSABRE default config.")

# ── tab:main36 ──
add_sheet("T_main36", "Table: 36-qubit suite (B-grid 2x2 4x4)",
    ["Circuit","CX","TS EPR","TS SWAP","dSABRE EPR","dSABRE SWAP","tket e-bits","Δ vs TS (%)","Δ vs tket (%)"],
    [["bv",17,5,34,1,8,3,-80.0,-66.7],
     ["dj",35,3,53,3,24,3,0.0,0.0],
     ["qaoa",1200,232,1027,145,1319,194,-37.5,-25.3],
     ["qpeexact",1019,100,771,65,874,175,-35.0,-62.9],
     ["vqe_su2",105,16,80,9,52,9,-43.8,0.0],
     ["wstate",70,12,88,8,63,6,-33.3,33.3],
     ["gmean","",20.1,147,11.3,95,16.0,-44.1,-29.4]])

# ── tab:main64 ──
add_sheet("T_main64", "Table: 64-qubit suite (H-grid 2x3 4x4, 96 physical qubits)",
    ["Circuit","CX","TS EPR","TS SWAP","dSABRE EPR","dSABRE SWAP","tket e-bits","Δ vs TS (%)","Δ vs tket (%)"],
    [["ae",1962,323,1508,216,2332,519,-33.1,-58.4],
     ["ghz",63,11,111,16,89,7,45.5,128.6],
     ["graphstate",64,76,379,19,104,9,-75.0,111.1],
     ["qft",1966,410,1872,246,2054,591,-40.0,-58.4],
     ["qnn",8126,1365,6292,521,8662,736,-61.8,-29.2],
     ["random*",1627,"","",714,3814,1181,"",-39.5],
     ["gmean (5)","",172.1,943,96.6,826,107.3,-43.9,-10.0],
     ["gmean (6)","","","",134.8,1066,160.0,"",-15.7]],
    notes="*TeleSABRE fails to converge on Random.")

# ── tab:ablation ──
add_sheet("T_ablation", "Table: Extended-set ablation (topological vs BFS)",
    ["Suite","dS-topo EPR","dS-topo SWAP","dSABRE EPR","dSABRE SWAP","Δ EPR (%)","Δ SWAP (%)"],
    [["25q",15.7,202,15.2,196,-3.2,-3.0],
     ["36q",12.1,88,11.3,95,-6.6,8.0],
     ["64q",151.5,1075,134.8,1066,-11.0,-0.8]])

# ── tab:mech ──
add_sheet("T_mech", "Table: Mechanism ablation (25q and 64q)",
    ["Configuration","25q gmEPR","25q Δ (%)","64q gmEPR","64q Δ (%)"],
    [["Full (baseline)",15.2,"---",134.8,"---"],
     ["No extended-set lookahead (w_e=0)",19.5,28.2,148.5,10.2],
     ["No capacity penalty (c_p=0)",24.3,59.9,213.9,58.7],
     ["No hop-gain reward (w_h=0)",15.2,0.0,137.4,1.9],
     ["No congestion relief",15.5,1.8,166.4,23.4]])

# ── tab:timing ──
add_sheet("T_timing", "Table: Compilation time in seconds (64-qubit suite)",
    ["Circuit","CX","TeleSABRE (s)","dSABRE (s)"],
    [["AE",1962,1.05,12.60],["GHZ",63,0.02,0.09],["Graphstate",64,3.92,0.05],
     ["QFT",1966,1.81,10.34],["QNN",8126,25.21,123.60],["Random*",1627,"",16.69]],
    notes="*TeleSABRE fails to converge on Random.")

# ── tab:large ──
add_sheet("T_large", "Table: Large-circuit scalability (QFT)",
    ["Suite","CX","TS EPR","TS SWAP","dSABRE EPR","dSABRE SWAP","tket e-bits","Δ vs TS (%)","Δ vs tket (%)"],
    [["100q",3420,248,2414,301,4077,815,21.4,-63.1],
     ["200q",7220,"","",1151,10989,1124,"",2.4],
     ["360q",13300,1109,15098,579,27489,669,-47.8,-13.5]],
    notes="TeleSABRE fails to converge on 200q QFT within 600s timeout.")

# ── tab:sensitivity ──
add_sheet("T_sensitivity", "Table: Cost-model sensitivity (gmean cost reduction of dSABRE vs TeleSABRE)",
    ["c_tele","25q (%)","36q (%)","64q (%)"],
    [[10,-20.6,-39.0,-23.8],
     [20,-25.2,-40.6,-29.0],
     [50,-31.5,-42.3,-35.4],
     [100,-35.2,-43.2,-38.9]],
    notes="c_swap=3 fixed; 64q excludes Random. Negative is better.")

# ── tab:node_decay_25 ──
add_sheet("T_nodedecay_25", "Table: Node-decay ablation, 25q suite",
    ["Circuit","EPR (on)","EPR (off)","Δ EPR (%)","LS (on)","LS (off)"],
    [["ae",23,23,0.0,374,304],["ghz",1,1,0.0,14,14],
     ["graphstate",2,2,0.0,16,14],["qft",33,33,0.0,461,349],
     ["qnn",48,48,0.0,1076,701],["random",169,178,5.3,1374,1123],
     ["gmean",15.2,15.3,0.9,"",""]])

# ── tab:node_decay_36 ──
add_sheet("T_nodedecay_36", "Table: Node-decay ablation, 36q suite",
    ["Circuit","EPR (on)","EPR (off)","Δ EPR (%)","LS (on)","LS (off)"],
    [["bv",1,1,0.0,8,9],["dj",3,3,0.0,24,22],
     ["qaoa",145,140,-3.4,1319,1006],["qpeexact",65,65,0.0,874,608],
     ["vqe_su2",9,15,66.7,52,59],["wstate",8,8,0.0,63,36],
     ["gmean",11.3,12.2,8.3,"",""]])

# ── tab:node_decay_64 ──
add_sheet("T_nodedecay_64", "Table: Node-decay ablation, 64q suite",
    ["Circuit","EPR (on)","EPR (off)","Δ EPR (%)","LS (on)","LS (off)"],
    [["ae",216,203,-6.0,2332,1363],["ghz",16,16,0.0,89,72],
     ["graphstate",19,15,-21.1,104,95],["qft",246,285,15.9,2054,1742],
     ["qnn",521,504,-3.3,8662,5167],["random",714,730,2.2,3814,3641],
     ["gmean",134.8,131.2,-2.7,"",""]])

# ── tab:hop_gain ──
add_sheet("T_hopgain", "Table: Hop-gain ablation across all three suites",
    ["Suite","Circuit","EPR (on)","EPR (off)","Δ EPR (%)"],
    [["25q","ae",23,23,0.0],["25q","ghz",1,1,0.0],["25q","graphstate",2,2,0.0],
     ["25q","qft",33,33,0.0],["25q","qnn",48,48,0.0],["25q","random",169,169,0.0],
     ["25q","gmean",15.2,15.2,0.0],
     ["36q","bv",1,1,0.0],["36q","dj",3,3,0.0],["36q","qaoa",145,145,0.0],
     ["36q","qpeexact",65,65,0.0],["36q","vqe_su2",9,9,0.0],["36q","wstate",8,8,0.0],
     ["36q","gmean",11.3,11.3,0.0],
     ["64q","ae",216,216,0.0],["64q","ghz",16,17,6.2],["64q","graphstate",19,19,0.0],
     ["64q","qft",246,246,0.0],["64q","qnn",521,521,0.0],["64q","random",714,751,5.2],
     ["64q","gmean",134.8,137.4,1.9]])

# ── tab:link_density ──
add_sheet("T_link_density", "Table: 4-link vs 8-link B-grid (2x2 4x4, 64 physical)",
    ["Suite","Circuit","pytket e-bits","TS 4L","TS 8L","dSABRE 4L","dSABRE 8L"],
    [["25q","ae",85,23,27,23,23],["25q","ghz",3,2,2,1,1],
     ["25q","graphstate",4,11,11,2,4],["25q","qft",120,39,37,33,33],
     ["25q","qnn",152,51,51,48,48],["25q","random",665,292,283,169,185],
     ["25q","gmean",48.1,25.8,26.1,15.2,17.3],
     ["25q","Δ 8L vs 4L (%)","---","+1.3","","+14",""],
     ["36q","bv",3,5,4,1,1],["36q","dj",3,3,5,3,3],
     ["36q","qaoa",194,232,254,145,139],["36q","qpeexact",175,100,207,65,68],
     ["36q","vqe_su2",9,16,15,9,12],["36q","wstate",6,12,12,8,9],
     ["36q","gmean",16.0,20.1,24.0,11.3,12.1],
     ["36q","Δ 8L vs 4L (%)","---","+19","","+7.0",""]],
    notes="pytket-dqc e-bit output is invariant under link multiplicity (single column).")

# ── tab:layout ──
add_sheet("T_layout", "Table: Initial-layout x routing-schedule gmean EPR",
    ["Layout","Schedule","25q EPR","64q EPR"],
    [["TS","fwd×2",22.5,110.3],["TS","sabre",19.3,98.4],
     ["Rand","fwd×2",32.0,125.4],["Rand","sabre",24.0,112.2],
     ["SL","fwd×2",17.6,112.3],["SL","sabre",15.7,100.2],
     ["TS reference","---",25.8,172.1]],
    notes="Bold-in-paper best: SL+sabre. Rand/SL best of 3 trials/seeds.")

# ── tab:pytket_shared_25_36 ──
add_sheet("T_pytket_25_36", "Table: Shared KaHyPar mapping — pytket-dqc vs dSABRE (25q, 36q)",
    ["Suite","Circuit","CX","pytket e-bits","pytket time (s)","dSABRE EPR","dSABRE time (s)","Δ"],
    [["25q","ae",558,85,1.0,23,0.9,"-73%"],
     ["25q","ghz",24,3,0.0,3,0.0,"0%"],
     ["25q","graphstate",25,4,0.0,5,0.0,"+25%"],
     ["25q","qft",580,120,1.2,33,0.8,"-72%"],
     ["25q","qnn",1223,152,3.7,49,2.7,"-68%"],
     ["25q","random",1124,665,3.0,181,5.5,"-73%"],
     ["25q","gmean","",48.1,"",21.6,"","-55%"],
     ["36q","bv",17,3,0.0,1,0.0,"-67%"],
     ["36q","dj",35,3,0.1,7,0.0,"+133%"],
     ["36q","qaoa",1200,194,3.8,156,3.2,"-20%"],
     ["36q","qpeexact",1019,175,2.3,87,2.2,"-50%"],
     ["36q","vqe_su2",105,9,0.2,9,0.1,"0%"],
     ["36q","wstate",70,6,0.1,8,0.0,"+33%"],
     ["36q","gmean","",16.0,"",13.8,"","-14%"]])

# ── tab:pytket_shared_64 ──
add_sheet("T_pytket_64", "Table: Shared KaHyPar mapping — pytket-dqc vs dSABRE (64q)",
    ["Circuit","CX","pytket e-bits","pytket time (s)","dSABRE EPR","dSABRE time (s)","Δ"],
    [["ae",1962,519,7.9,226,11.2,"-56%"],
     ["ghz",63,7,0.1,7,0.0,"0%"],
     ["graphstate",64,9,0.2,8,0.0,"-11%"],
     ["qft",1966,591,9.6,229,9.2,"-61%"],
     ["qnn",8126,736,131.6,518,134.9,"-30%"],
     ["random",1627,1181,6.4,757,16.5,"-36%"],
     ["gmean","",160.0,"",102.2,"","-36%"]])

# ── tab:dmaps_results ──
add_sheet("T_dmaps_results", "Table: DMapS results on B-grid (25q, 36q) and H-grid (64q)",
    ["Suite","Circuit","EPR","lSWAP","rCX","rSWAP","best","mean","stdev","range","time (s)"],
    [["25q","ghz",1,4,1,0,14,15.6,1.5,"14-18",4.7],
     ["25q","graphstate",2,5,2,0,25,29.6,5.2,"25-38",4.9],
     ["25q","qft",121,231,49,36,1441,1620.0,123.5,"1441-1764",16.7],
     ["25q","qnn",121,504,51,35,1714,1783.8,82.7,"1714-1924",26.5],
     ["25q","random",342,1119,112,115,4539,4682.6,131.8,"4539-4846",56.7],
     ["25q","ae†",80,223,40,20,1023,1199.8,130.0,"1023-1317",16.2],
     ["36q","bv†",7,23,5,1,93,94.8,1.3,"93-96",5.1],
     ["36q","dj",28,115,18,5,395,429.2,28.0,"395-465",6.2],
     ["36q","qaoa",466,958,166,150,5618,6742.6,1053.1,"5618-8277",62.6],
     ["36q","qpeexact†",182,446,76,53,2266,2485.0,194.2,"2266-2760",26.9],
     ["36q","vqe_su2",6,8,6,0,68,72.0,4.3,"68-78",6.8],
     ["36q","wstate",4,2,4,0,42,46.0,4.8,"42-53",6.1],
     ["64q","ghz",3,17,3,0,47,50.4,2.4,"47-53",8.7],
     ["64q","graphstate",4,24,4,0,64,66.8,3.1,"64-72",8.9],
     ["64q","qft",392,1152,148,122,5072,5350.6,247.4,"5072-5648",74.2],
     ["64q","qnn",933,3762,405,264,13092,15422.6,1666.4,"13092-17035",306.2],
     ["64q","random",1138,3446,378,380,14826,15132.8,278.4,"14826-15555",150.9],
     ["64q","ae†",346,1119,116,115,4579,5167.8,734.2,"4579-6359",67.1]],
    notes="Best of 5 KaHyPar seeds. EPR=rCX+2*rSWAP. Overall uses DMapS weighting 1:10:20. † originally failed; runs after upstream patches.")

# ── tab:dmaps_vs_dsabre ──
add_sheet("T_dmaps_vs_dsabre", "Table: DMapS vs dSABRE head-to-head",
    ["Suite","Circuit","CX","dSABRE EPR","dSABRE lSW","rCX","rSWAP","DMapS EPR","DMapS lSW","EPR×","lSW×"],
    [["25q","ghz",24,1,14,1,0,1,4,1.00,0.29],
     ["25q","graphstate",25,2,16,2,0,2,5,1.00,0.31],
     ["25q","qft",580,33,461,49,36,121,231,3.67,0.50],
     ["25q","qnn",1223,48,1076,51,35,121,504,2.52,0.47],
     ["25q","random",1124,169,1374,112,115,342,1119,2.02,0.81],
     ["25q","ae†",558,23,374,40,20,80,223,3.48,0.60],
     ["36q","bv†",17,1,8,5,1,7,23,7.00,2.88],
     ["36q","dj",35,3,24,18,5,28,115,9.33,4.79],
     ["36q","qaoa",1200,145,1319,166,150,466,958,3.21,0.73],
     ["36q","qpeexact†",1019,65,874,76,53,182,446,2.80,0.51],
     ["36q","vqe_su2",105,9,52,6,0,6,8,0.67,0.15],
     ["36q","wstate",70,8,63,4,0,4,2,0.50,0.03],
     ["64q","ghz",63,16,89,3,0,3,17,0.19,0.19],
     ["64q","graphstate",64,19,104,4,0,4,24,0.21,0.23],
     ["64q","qft",1966,246,2054,148,122,392,1152,1.59,0.56],
     ["64q","qnn",8126,521,8662,405,264,933,3762,1.79,0.43],
     ["64q","random",1627,714,3814,378,380,1138,3446,1.59,0.90],
     ["64q","ae†",1962,216,2332,116,115,346,1119,1.60,0.48]],
    notes="DMapS best-of-5; dSABRE best-of-3 SabreLayout seeds. Values <1 mean DMapS wins.")

# ── tab:dmaps_fill ──
add_sheet("T_dmaps_fill", "Table: Fill-ratio sweep — QFT on fixed 4-chip B-grid (64 phys)",
    ["n","fill","CX","dSABRE EPR","dSABRE lSW","DMapS EPR","DMapS lSW","EPR×","lSW×"],
    [[24,"37.5%",588,30,446,118,238,3.93,0.53],
     [36,"56.2%",1314,88,1153,406,803,4.61,0.70],
     [48,"75.0%",2272,159,2084,518,1301,3.26,0.62],
     [56,"87.5%",2924,819,4423,831,1523,1.01,0.34]])

# Reorder: Index first
wb._sheets = [wb["Index"]] + [s for s in wb._sheets if s.title != "Index"]

out = "/Users/sanjiangli/Documents/pyzoo/dsabre/paper/paper_tables.xlsx"
wb.save(out)
print(f"Wrote {out} with {len(wb.sheetnames)} sheets")
